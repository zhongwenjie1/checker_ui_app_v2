# -*- coding: utf-8 -*-
from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFileDialog, QToolBar, QStatusBar, QMessageBox, QTableWidget,
    QTableWidgetItem, QSpinBox, QComboBox, QLineEdit, QColorDialog,
    QGraphicsScene, QGraphicsView, QGraphicsRectItem, QGraphicsTextItem, QGraphicsLineItem, QSplitter, QDialog, QTabWidget
)
from PySide6.QtCore import Qt, QThreadPool, QPointF
from PySide6.QtGui import QAction, QColor, QPainter, QPainterPath, QPen, QBrush
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Border, Side

# 版本号：优先从当前工程的 __init__ 里取，取不到就用 "dev"
try:
    from __init__ import __version__
except Exception:
    __version__ = "dev"

# Worker & tickets：直接从当前工程内部模块导入
from infra.threads import Worker
from core import tickets


class ExportTicketWindow(QMainWindow):
    """
    导出组合票（独立于数据校对）
    步骤表字段：
      序号 / 工序显示名 / 工位组名 / 并行能力 / 时长(秒，可逗号) /
      区域ID(可选) / 区域容量(可选) / 起始需等区域ID(可选)
    说明：
      - 同一“区域ID”的一串连续步骤视为一个“阻塞区域（Zone）”，容量=同时允许几台车处于该区域。
      - “起始需等区域ID”用于上游工位：本工位本身不占用区域名额，但开工/放行必须等待该区域出现空位。
    """
    COL_COLOR = 8
    MAX_SINGLE_STEPS = 40  # 单工程组合票：模板最多支持的行数

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"导出组合票  v{__version__}")
        self.resize(1120, 720)

        self.thread_pool = QThreadPool.globalInstance()
        self.dst_path = None

        self._build_ui()
        self._connect_signals()
    def _on_tab_changed(self, index: int):
        """
        Tab 切换时，控制顶部“添加步骤 / 删除步骤 / 填入示例”按钮：
        - 仅在『多工程组合票』页签（第 0 个 Tab）启用；
        - 在『单工程组合票』页签禁用，避免误点影响单工程表。
        """
        is_multi = (index == 0)
        self.act_add_row.setEnabled(is_multi)
        self.act_del_row.setEnabled(is_multi)
        self.act_fill_sample.setEnabled(is_multi)
    # ---------------- UI ---------------- #
    def _build_ui(self):
        tb = QToolBar("Ticket")
        self.addToolBar(tb)

        self.act_back = QAction("返回主页", self)
        tb.addAction(self.act_back)
        tb.addSeparator()

        self.act_help = QAction("帮助", self)
        tb.addAction(self.act_help)
        tb.addSeparator()

        self.act_diagram = QAction("流程图", self)
        tb.addAction(self.act_diagram)
        tb.addSeparator()

        self.act_add_row = QAction("添加步骤", self)
        self.act_del_row = QAction("删除步骤", self)
        self.act_fill_sample = QAction("填入示例", self)
        tb.addAction(self.act_add_row)
        tb.addAction(self.act_del_row)
        tb.addSeparator()
        tb.addAction(self.act_fill_sample)

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        # ====== Tab 控件 ======
        self.tabs = QTabWidget(self)
        root.addWidget(self.tabs)

        # ---------- Tab1：多车组合票 ----------
        self.page_multi = QWidget(self)
        page_multi_layout = QVBoxLayout(self.page_multi)
        self.tabs.addTab(self.page_multi, "多工程组合票")

        # 参数区
        row_top = QHBoxLayout()
        page_multi_layout.addLayout(row_top)

        row_top.addWidget(QLabel("工程名称："))
        self.ed_project = QLineEdit()
        self.ed_project.setPlaceholderText("例如：L2++")
        self.ed_project.setFixedWidth(220)
        row_top.addWidget(self.ed_project)

        row_top.addSpacing(12)
        row_top.addWidget(QLabel("车号数量："))
        self.spn_cars = QSpinBox()
        self.spn_cars.setRange(1, 9999)
        self.spn_cars.setValue(4)
        row_top.addWidget(self.spn_cars)

        row_top.addSpacing(12)
        row_top.addWidget(QLabel("时间格刻度："))
        self.cmb_grid = QComboBox()
        self.cmb_grid.addItems(["1.0", "0.5", "2.0"])
        self.cmb_grid.setCurrentIndex(0)
        row_top.addWidget(self.cmb_grid)

        row_top.addSpacing(12)
        row_top.addWidget(QLabel("等待分配："))
        self.cmb_wait = QComboBox()
        self.cmb_wait.addItems(["开始前等待", "末尾等待"])
        self.cmb_wait.setCurrentIndex(0)
        row_top.addWidget(self.cmb_wait)

        row_top.addStretch()

        # 步骤表：新增“起始需等区域ID(可选)”和“填充颜色(可选)”
        self.tbl = QTableWidget(0, 9, self)
        self.tbl.setHorizontalHeaderLabels([
            "序号", "工序显示名", "工位组名", "并行能力",
            "时长(秒，逗号分隔表示多台)", "区域ID(可选)", "区域容量(可选)",
            "起始需等区域ID(可选)", "填充颜色(可选)"
        ])
        self.tbl.horizontalHeader().setStretchLastSection(True)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setColumnWidth(self.COL_COLOR, 40)

        page_multi_layout.addWidget(self.tbl, 1)

        # 底部导出按钮栏
        btn_bar = QHBoxLayout()
        page_multi_layout.addLayout(btn_bar)
        btn_bar.addStretch()
        self.btn_export = QPushButton("生成并导出组合票")
        btn_bar.addWidget(self.btn_export)

        # ---------- Tab2：单工程组合票 ----------
        self.page_single = QWidget(self)
        page_single_layout = QVBoxLayout(self.page_single)
        page_single_layout.setContentsMargins(8, 8, 8, 8)
        page_single_layout.setSpacing(8)

        # 顶部基本信息
        row_info = QHBoxLayout()
        page_single_layout.addLayout(row_info)

        row_info.addWidget(QLabel("工程名称："))
        self.ed_sw_project = QLineEdit(self.page_single)
        self.ed_sw_project.setPlaceholderText("例如：前轴调整工位")
        self.ed_sw_project.setFixedWidth(200)
        row_info.addWidget(self.ed_sw_project)

        row_info.addSpacing(12)
        row_info.addWidget(QLabel("品番·品名："))
        self.ed_sw_part = QLineEdit(self.page_single)
        self.ed_sw_part.setPlaceholderText("例如：XXXX-XXXXX 前轮定位")
        self.ed_sw_part.setFixedWidth(220)
        row_info.addWidget(self.ed_sw_part)

        row_info.addSpacing(12)
        row_info.addWidget(QLabel("作业者："))
        self.ed_sw_worker = QLineEdit(self.page_single)
        self.ed_sw_worker.setPlaceholderText("例如：张三")
        self.ed_sw_worker.setFixedWidth(120)
        row_info.addWidget(self.ed_sw_worker)

        row_info.addSpacing(12)
        row_info.addWidget(QLabel("节拍TT(秒)："))
        self.spn_sw_takt = QSpinBox(self.page_single)
        self.spn_sw_takt.setRange(1, 9999)
        self.spn_sw_takt.setValue(118)  # 默认示例
        row_info.addWidget(self.spn_sw_takt)

        row_info.addStretch()

        # 作业手顺表（A→B 区间）
        self.tbl_sw = QTableWidget(0, 7, self.page_single)
        self.tbl_sw.setHorizontalHeaderLabels([
            "顺序", "作业名称A", "作业名称B", "手作业(秒)", "自动(秒)", "步行(秒)", "步行在前/后"
        ])
        self.tbl_sw.horizontalHeader().setStretchLastSection(True)
        self.tbl_sw.verticalHeader().setVisible(False)
        page_single_layout.addWidget(self.tbl_sw, 1)

        # 底部按钮栏（单工程组合票）
        row_btn_sw = QHBoxLayout()
        page_single_layout.addLayout(row_btn_sw)
        row_btn_sw.addStretch()

        self.btn_sw_add = QPushButton("添加作业行", self.page_single)
        self.btn_sw_del = QPushButton("删除选中行", self.page_single)
        self.btn_sw_export = QPushButton("导出标准作业组合票", self.page_single)

        row_btn_sw.addWidget(self.btn_sw_add)
        row_btn_sw.addWidget(self.btn_sw_del)
        row_btn_sw.addWidget(self.btn_sw_export)

        self.tabs.addTab(self.page_single, "单工程组合票")

        # 状态栏（用于显示导出进度 / 完成信息）
        self.status = QStatusBar()
        self.setStatusBar(self.status)

    def _connect_signals(self):
        self.act_back.triggered.connect(self.go_home)
        self.act_add_row.triggered.connect(self.add_row)
        self.act_del_row.triggered.connect(self.del_row)
        self.act_fill_sample.triggered.connect(self.fill_sample)
        self.btn_export.clicked.connect(self.do_export)
        self.act_help.triggered.connect(self.show_help)
        self.act_diagram.triggered.connect(self.show_diagram)

        # Tab 切换时，控制顶部步骤按钮是否可用
        self.tabs.currentChanged.connect(self._on_tab_changed)

        # 单工程组合票 Tab
        self.btn_sw_add.clicked.connect(self.add_single_row)
        self.btn_sw_del.clicked.connect(self.del_single_row)
        self.btn_sw_export.clicked.connect(self.export_single_placeholder)

    # ------------- 多车组合票：动作 ------------- #
    def add_row(self):
        r = self.tbl.rowCount()
        self.tbl.insertRow(r)
        # 默认值：序号递增、能力=1、区域留空
        self.tbl.setItem(r, 0, QTableWidgetItem(str(r + 1)))
        self.tbl.setItem(r, 1, QTableWidgetItem(""))
        self.tbl.setItem(r, 2, QTableWidgetItem(""))
        self.tbl.setItem(r, 3, QTableWidgetItem("1"))
        self.tbl.setItem(r, 4, QTableWidgetItem(""))
        self.tbl.setItem(r, 5, QTableWidgetItem(""))   # 区域ID(可选)
        self.tbl.setItem(r, 6, QTableWidgetItem(""))   # 区域容量(可选)
        self.tbl.setItem(r, 7, QTableWidgetItem(""))   # 起始需等区域ID(可选)

        color_btn = QPushButton("…")
        color_btn.setFixedSize(30, 22)
        color_btn.clicked.connect(lambda _, row=r: self._choose_color(row))
        self.tbl.setCellWidget(r, self.COL_COLOR, color_btn)
        color_item = QTableWidgetItem("")
        color_item.setData(Qt.UserRole, "")
        self.tbl.setItem(r, self.COL_COLOR, color_item)

    def _choose_color(self, row: int):
        dlg_col = QColorDialog.getColor(parent=self)
        if dlg_col.isValid():
            hex_code = dlg_col.name()
            btn = self.tbl.cellWidget(row, self.COL_COLOR)
            btn.setStyleSheet(f"background:{hex_code};")
            self.tbl.item(row, self.COL_COLOR).setData(Qt.UserRole, hex_code)

    def del_row(self):
        r = self.tbl.currentRow()
        if r >= 0:
            self.tbl.removeRow(r)

    # -------- 单人标准作业组合票：行操作 --------
    def add_single_row(self):
        """在单人作业手顺表中新增一行"""
        if not hasattr(self, "tbl_sw"):
            return
        r = self.tbl_sw.rowCount()
        self.tbl_sw.insertRow(r)
        # 顺序默认递增（组合票行号）
        self.tbl_sw.setItem(r, 0, QTableWidgetItem(str(r + 1)))
        # 作业名称A / B 先留空，让你填写
        self.tbl_sw.setItem(r, 1, QTableWidgetItem(""))
        self.tbl_sw.setItem(r, 2, QTableWidgetItem(""))
        # 手作业 / 自动 / 步行，默认 0
        self.tbl_sw.setItem(r, 3, QTableWidgetItem("0"))
        self.tbl_sw.setItem(r, 4, QTableWidgetItem("0"))
        self.tbl_sw.setItem(r, 5, QTableWidgetItem("0"))
        # 步行位置：默认“后置”
        pos_cb = QComboBox(self.tbl_sw)
        pos_cb.addItem("后置", userData="after")
        pos_cb.addItem("前置", userData="before")
        self.tbl_sw.setCellWidget(r, 6, pos_cb)

    def del_single_row(self):
        """删除单人作业手顺表中的选中行"""
        if not hasattr(self, "tbl_sw"):
            return
        r = self.tbl_sw.currentRow()
        if r >= 0:
            self.tbl_sw.removeRow(r)
        # 重写顺序列，让它保持 1,2,3,...
        for i in range(self.tbl_sw.rowCount()):
            item = self.tbl_sw.item(i, 0)
            if item is None:
                item = QTableWidgetItem()
                self.tbl_sw.setItem(i, 0, item)
            item.setText(str(i + 1))

    # -------- 单人标准作业组合票：数据收集 --------
    def _collect_single_inputs(self):
        """
        从单人作业手顺 Tab 中读取数据，并计算时间汇总。
        返回：
          project, part, worker, takt_sec, steps, totals
        其中：
          steps: [{seq, name, name_a, name_b, manual, auto, walk, walk_pos, duration, start, end}, ...]
          totals: {"manual": x, "auto": y, "walk": z, "total": t}
        """
        if not hasattr(self, "tbl_sw"):
            raise ValueError("单人作业手顺表尚未初始化")

        project = (self.ed_sw_project.text().strip() or "工程")
        part = self.ed_sw_part.text().strip()
        worker = self.ed_sw_worker.text().strip()
        takt_sec = int(self.spn_sw_takt.value())

        steps = []
        cur_time = 0.0
        total_manual = 0.0
        total_auto = 0.0
        total_walk = 0.0

        for r in range(self.tbl_sw.rowCount()):
            # 作业名称 A / B
            name_a_item = self.tbl_sw.item(r, 1)
            name_b_item = self.tbl_sw.item(r, 2)
            name_a = name_a_item.text().strip() if name_a_item else ""
            name_b = name_b_item.text().strip() if name_b_item else ""

            if not name_a and not name_b:
                # 两个都没填，当作空行，跳过
                continue

            # 导出时使用的显示名（A→B / 单独一个）
            if name_a and name_b:
                name = f"{name_a} → {name_b}"
            else:
                name = name_a or name_b

            def _get_time(col_idx: int) -> float:
                item = self.tbl_sw.item(r, col_idx)
                txt = item.text().strip() if item else ""
                if not txt:
                    return 0.0
                try:
                    return float(txt)
                except Exception:
                    raise ValueError(f"第 {r + 1} 行时间列（第 {col_idx + 1} 列）不是有效数字：{txt}")

            # 手作业 / 自动 / 步行时间列：3, 4, 5
            manual = _get_time(3)
            auto = _get_time(4)
            walk = _get_time(5)

            # 步行位置：前置/后置（默认后置）
            walk_pos = "after"
            pos_widget = self.tbl_sw.cellWidget(r, 6)
            if isinstance(pos_widget, QComboBox):
                walk_pos_data = pos_widget.currentData()
                if walk_pos_data in ("before", "after"):
                    walk_pos = walk_pos_data

            duration = manual + auto + walk
            if duration <= 0:
                raise ValueError(f"第 {r + 1} 行『{name}』的时间合计为 0，请填写手作业/自动/步行时间。")

            start = cur_time
            end = cur_time + duration
            cur_time = end

            total_manual += manual
            total_auto += auto
            total_walk += walk

            # 顺序列（如果用户改过，我们尽量读取）
            seq_item = self.tbl_sw.item(r, 0)
            try:
                seq = int(seq_item.text()) if seq_item and seq_item.text().strip() else len(steps) + 1
            except Exception:
                seq = len(steps) + 1

            steps.append({
                "seq": seq,
                "name": name,       # A→B 组合显示名（保留）
                "name_a": name_a,   # 原始作业名称A
                "name_b": name_b,   # 原始作业名称B
                "manual": manual,
                "auto": auto,
                "walk": walk,
                "walk_pos": walk_pos,  # 步行在前/后
                "duration": duration,
                "start": start,
                "end": end,
            })

        # 行数上限检查：防止超过模板预留的行数
        if len(steps) > self.MAX_SINGLE_STEPS:
            raise ValueError(
                f"当前单人标准作业组合票共有 {len(steps)} 行，已超过模板最多支持的 {self.MAX_SINGLE_STEPS} 行。\n"
                "请合并部分区间或拆分为多张组合票后再导出。"
            )

        if not steps:
            raise ValueError("请至少填写一行有效的作业（需有作业名称和时间）。")

        totals = {
            "manual": total_manual,
            "auto": total_auto,
            "walk": total_walk,
            "total": total_manual + total_auto + total_walk,
        }
        return project, part, worker, takt_sec, steps, totals

    # -------- 单人标准作业组合票：写入模板 --------
    def _export_single_to_excel(self, path, project, part, worker, takt_sec, steps, totals):
        """
        根据单人作业手顺（A→B 区间）将数据写入《组合票标准版.xlsx》模板：
        - 模板文件需放在与本文件同一目录下，文件名：组合票标准版.xlsx
        - 仅填充左侧步骤表区域和基本信息，不修改模板中的其他格式/图表
        """
        # 定位模板文件：与本 .py 同目录
        base_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(base_dir, "组合票标准版.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"未找到模板文件：{template_path}")

        wb = load_workbook(template_path)
        try:
            ws = wb["④标准作业组合票"]
        except KeyError:
            ws = wb.active

        def _set_value(coord, value):
            """安全写入单元格：若目标是合并单元格，从其合并区域左上角写入"""
            cell = ws[coord]
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        ws.cell(row=mr.min_row, column=mr.min_col).value = value
                        break
            else:
                cell.value = value

        def _set_fill(row, col, fill):
            """安全设置单元格填充：若目标是合并单元格，则写到其合并区域左上角"""
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        ws.cell(row=mr.min_row, column=mr.min_col).fill = fill
                        break
            else:
                cell.fill = fill

        def _set_border(row, col, border: Border):
            """
            安全设置单元格边框：若目标是合并单元格，则写到其合并区域左上角；
            与已有边框合并（只改指定方向的线型）。
            """
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        cell = ws.cell(row=mr.min_row, column=mr.min_col)
                        break

            old = cell.border or Border()

            def merge_side(new_side, old_side):
                if getattr(new_side, "style", None):
                    return new_side
                return old_side

            cell.border = Border(
                left=merge_side(border.left, old.left),
                right=merge_side(border.right, old.right),
                top=merge_side(border.top, old.top),
                bottom=merge_side(border.bottom, old.bottom),
                diagonal=old.diagonal,
                diagonal_direction=old.diagonal_direction,
                outline=old.outline,
                vertical=old.vertical,
                horizontal=old.horizontal,
            )

        # 1) 清空左侧原有数据区域
        start_row = 9
        row_span = 3
        max_steps = getattr(self, "MAX_SINGLE_STEPS", 40)
        end_row = start_row + max_steps * row_span - 1
        for row in range(start_row, end_row + 1):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        # 清空右侧时间轴区域填充（F列开始，按总时间估算范围）
        time_start_col = 6  # F列
        max_time = int(round(totals.get("total", 0))) if isinstance(totals, dict) else 0
        if max_time < 0:
            max_time = 0
        time_end_col = time_start_col + max_time + 5
        for row in range(start_row, end_row + 1):
            for col in range(time_start_col, time_end_col + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.fill = PatternFill()

        # 2) 写入步骤：每步占 3 行（A9:A11, A12:A14, ...）
        row_span = 3
        time_start_col = 6  # F列
        time_fill = PatternFill(fill_type="solid", fgColor="000000")
        bars = []  # 记录每个黑条，用于画连线

        for idx, s in enumerate(steps):
            base_row = start_row + idx * row_span

            # 序号
            ws.cell(row=base_row, column=1).value = s["seq"]

            # 作业名称 A/B：B 列两行
            name_a = s.get("name_a") or s.get("name") or ""
            name_b = s.get("name_b") or ""
            ws.cell(row=base_row, column=2).value = name_a
            if name_b:
                ws.cell(row=base_row + 2, column=2).value = name_b

            # 时间数值（C~E）
            ws.cell(row=base_row, column=3).value = s["manual"]
            ws.cell(row=base_row, column=4).value = s["auto"]
            ws.cell(row=base_row, column=5).value = s["walk"]

            # 时间条（手+自，步行空白）
            start_sec = int(round(s["start"]))
            manual_auto = float(s["manual"]) + float(s["auto"])
            walk = float(s["walk"])
            walk_pos = s.get("walk_pos", "after")

            if walk_pos == "before":
                bar_start_sec = int(round(start_sec + walk))
            else:
                bar_start_sec = start_sec

            bar_end_sec = int(round(bar_start_sec + manual_auto))

            if bar_end_sec > bar_start_sec:
                mid_row = base_row + 1
                for sec in range(bar_start_sec, bar_end_sec):
                    col = time_start_col + sec
                    _set_fill(mid_row, col, time_fill)

                bars.append(
                    {
                        "mid_row": mid_row,
                        "bar_start": bar_start_sec,
                        "bar_end": bar_end_sec,
                    }
                )

        # 2.5) 相邻黑条之间画连接线：
        #      - 有间隔：步行 → 虚折线，从黑条末端下端开始，先竖后横（形状参考样本图）
        #      - 无间隔：直接接续 → 加粗实直线
        if len(bars) >= 2:
            # 线条加粗一些，效果更明显
            solid_side = Side(style="medium", color="000000")        # 加粗实线
            walk_side = Side(style="mediumDashed", color="000000")   # 加粗虚线

            h_walk_border = Border(top=walk_side)    # 水平虚线
            v_walk_border = Border(left=walk_side)   # 垂直虚线
            v_solid_border = Border(left=solid_side) # 垂直实线

            for i in range(len(bars) - 1):
                curr = bars[i]
                nxt = bars[i + 1]

                mid_row_curr = curr["mid_row"]     # 当前黑条所在行（块中间行）
                mid_row_nxt = nxt["mid_row"]       # 下一黑条所在行
                bar_end_curr = curr["bar_end"]
                bar_start_nxt = nxt["bar_start"]

                # 黑条最后一格所在的列（注意 bar_end 是“结束时间”，最后一格是 bar_end-1）
                col_end_prev = time_start_col + bar_end_curr - 1
                col_start_next = time_start_col + bar_start_nxt

                if bar_start_nxt > bar_end_curr:
                    # 有间隔：步行 → 虚折线
                    # 右侧连接列：在黑条的“右边一格”开始
                    col_conn = time_start_col + bar_end_curr       # 黑条最后一格的右侧列
                    col_start_next = time_start_col + bar_start_nxt

                    # 1) 竖虚线：从当前黑条的下沿(mid_row_curr+1) 到「下一条黑条的上一行(mid_row_nxt-1)」
                    row_vert_start = mid_row_curr + 1
                    row_vert_end = mid_row_nxt - 1
                    if row_vert_start <= row_vert_end:
                        for row in range(row_vert_start, row_vert_end + 1):
                            _set_border(row, col_conn, v_walk_border)

                    # 2) 水平虚线：在下一条黑条所在行，从竖线落点一直画到下一条黑条起点前一格
                    #    即 [col_conn, col_start_next-1]，注意 range 右开区间
                    for col in range(col_conn, col_start_next):
                        _set_border(mid_row_nxt, col, h_walk_border)
                else:
                    # 无间隔或略微重叠：视为直接接续 → 加粗实直线
                    # 连接列选在“前一条最后一格”和“下一条开始”的较大值所在列
                    col = time_start_col + max(bar_end_curr - 1, bar_start_nxt)
                    row_top = min(mid_row_curr, mid_row_nxt)
                    row_bottom = max(mid_row_curr, mid_row_nxt)
                    for row in range(row_top, row_bottom + 1):
                        _set_border(row, col, v_solid_border)

        # 3) 合计行：B49 总时间，C49 手作业总时间，D49 自动总时间，E49 步行总时间
        if isinstance(totals, dict):
            total_sec = totals.get("total", 0.0)
            manual_sec = totals.get("manual", 0.0)
            auto_sec = totals.get("auto", 0.0)
            walk_sec = totals.get("walk", 0.0)
        else:
            total_sec = manual_sec = auto_sec = walk_sec = 0.0

        def _fmt_sec(v):
            """把秒数统一转成整数秒写入单元格"""
            try:
                return int(round(float(v)))
            except Exception:
                return v

        _set_value("B49", _fmt_sec(total_sec))   # 合计下面：总时间
        _set_value("C49", _fmt_sec(manual_sec))  # 手作业合计
        _set_value("D49", _fmt_sec(auto_sec))    # 自动合计
        _set_value("E49", _fmt_sec(walk_sec))    # 步行合计

        # 4) 在上方空白处写入工程信息
        _set_value("B2", project)
        _set_value("B3", part)
        _set_value("B4", worker)
        _set_value("E2", takt_sec)

        # 5) 保存
        wb.save(path)

    def export_single_placeholder(self):
        """
        单工程组合票导出流程：
        1. 读取 Tab2 中 A→B 区间作业数据并校验
        2. 选择保存路径
        3. 使用固定 Excel 模板导出标准作业组合票
        """
        try:
            project, part, worker, takt_sec, steps, totals = self._collect_single_inputs()
        except Exception as e:
            QMessageBox.warning(self, "输入有误", str(e))
            return

        default_name = f"{project}_单人组合票.xlsx" if project else "单人组合票.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出标准作业组合票",
            default_name,
            "Excel (*.xlsx)",
        )
        if not path:
            return

        try:
            self._export_single_to_excel(path, project, part, worker, takt_sec, steps, totals)
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))
            return

        msg = (
            f"已导出标准作业组合票：\n{path}\n\n"
            f"工程名称：{project}\n"
            f"品番·品名：{part or '（未填写）'}\n"
            f"作业者：{worker or '（未填写）'}\n\n"
            f"节拍 TT：{takt_sec} 秒\n"
            f"总时间：{totals['total']:.1f} 秒\n"
            f"  其中 手作业：{totals['manual']:.1f} 秒\n"
            f"       自动：{totals['auto']:.1f} 秒\n"
            f"       步行：{totals['walk']:.1f} 秒\n\n"
            f"步骤数：{len(steps)} 步"
        )
        QMessageBox.information(self, "单工程组合票 - 导出完成", msg)

    # -------- 多车组合票：数据收集 & 导出 --------
    def fill_sample(self):
        """
        串行示例 + 阻塞区域 + 上游闸门 示例
        """
        self.tbl.setRowCount(0)
        sample_rows = [
            # 序号, 显示名,   组,       能力, 时长,   区域ID, 容量, 起始需等区域
            ("1",  "L2++",   "L2++",   "1", "112", "",   "",   "Z1"),
            ("2",  "电检准备", "电检准备", "1", "39.5", "",   "",   "Z1"),
            ("3",  "电检1",   "电检",   "1", "80",  "",   "",   ""),   # 普通工位
            ("4",  "电检2",   "电检",   "1", "70",  "Z1", "1", ""),   # 区域入口
            ("5",  "电检结束", "电检结束", "1", "29.5","Z1", "",  ""),   # 区域内
            ("6",  "NDA圈外", "NDA外",   "1", "30",  "Z1", "",  ""),   # 区域内
            ("7",  "NDA圈内", "NDA内",   "1", "20",  "Z1", "",  ""),   # 区域出口
            ("8",  "NDA检查", "NDA检查", "1", "30",  "",   "",   ""),   # 区域外
        ]
        for row in sample_rows:
            self.add_row()
            r = self.tbl.rowCount() - 1
            for c, text in enumerate(row):
                self.tbl.setItem(r, c, QTableWidgetItem(str(text)))

        if not self.ed_project.text().strip():
            self.ed_project.setText("L2++")
        self.spn_cars.setValue(4)
        self.cmb_grid.setCurrentText("1.0")
        self.cmb_wait.setCurrentText("开始前等待")

    def _collect_inputs(self):
        project = self.ed_project.text().strip() or "工程"
        cars = int(self.spn_cars.value())
        try:
            grid_step = float(self.cmb_grid.currentText())
            if grid_step <= 0:
                grid_step = 1.0
        except Exception:
            grid_step = 1.0
        wait_policy = "before" if self.cmb_wait.currentIndex() == 0 else "after"

        defs = []
        for r in range(self.tbl.rowCount()):
            seq = (self.tbl.item(r, 0).text().strip() if self.tbl.item(r, 0) else "")
            name = (self.tbl.item(r, 1).text().strip() if self.tbl.item(r, 1) else "")
            grp = (self.tbl.item(r, 2).text().strip() if self.tbl.item(r, 2) else "")
            cap = (self.tbl.item(r, 3).text().strip() if self.tbl.item(r, 3) else "1")
            dur = (self.tbl.item(r, 4).text().strip() if self.tbl.item(r, 4) else "")
            zid = (self.tbl.item(r, 5).text().strip() if self.tbl.item(r, 5) else "")
            zcap = (self.tbl.item(r, 6).text().strip() if self.tbl.item(r, 6) else "")
            gzd = (self.tbl.item(r, 7).text().strip() if self.tbl.item(r, 7) else "")
            color_hex = self.tbl.item(r, self.COL_COLOR).data(Qt.UserRole) or ""

            if not name or not grp or not dur:
                continue

            try:
                capacity = max(1, int(float(cap)))
            except Exception:
                capacity = 1

            durations = []
            for part in dur.replace("，", ",").split(","):
                t = part.strip()
                if not t:
                    continue
                try:
                    durations.append(float(t))
                except Exception:
                    pass
            if not durations:
                continue

            rec = {
                "seq": int(float(seq)) if seq else len(defs) + 1,
                "display": name,
                "group": grp,
                "capacity": capacity,
                "durations": durations,
                "color": color_hex,
            }

            if zid:
                rec["zone_id"] = zid
                try:
                    rec["zone_capacity"] = max(1, int(float(zcap))) if zcap else 1
                except Exception:
                    rec["zone_capacity"] = 1

            if gzd:
                rec["gate_zone_id"] = gzd

            defs.append(rec)

        defs.sort(key=lambda x: x["seq"])
        if not defs:
            raise ValueError("请至少填写一行有效的步骤（工序显示名/工位组名/时长）")

        return project, cars, grid_step, wait_policy, defs

    def do_export(self):
        try:
            project, cars, grid_step, wait_policy, defs = self._collect_inputs()
        except Exception as e:
            QMessageBox.warning(self, "输入有误", str(e))
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出位置",
            f"{project}_组合票.xlsx",
            "Excel (*.xlsx)",
        )
        if not path:
            return
        self.dst_path = path

        worker = Worker(
            tickets.schedule_and_export,
            defs, cars, grid_step, wait_policy, project, self.dst_path,
        )
        worker.signals.error.connect(self._on_error)
        worker.signals.finished.connect(self._on_export_finished)
        self.thread_pool.start(worker)
        self.status.showMessage("正在生成组合票...", 5000)

    def _on_export_finished(self):
        self.status.showMessage("导出完成", 6000)
        QMessageBox.information(self, "完成", f"已导出：\n{self.dst_path}")

    def go_home(self):
        home = getattr(self, "home_window", None)
        if home is not None and hasattr(home, "show"):
            try:
                home.show()
            except Exception:
                pass
        self.close()

    # ---------- 帮助弹窗 ----------
    def show_help(self):
        msg = (
            "<h3>组合票操作指南</h3>"
            "<ol>"
            "<li>点击『添加步骤』逐行录入；列含 ‘区域ID/容量’ 与 ‘起始需等区域ID’</li>"
            "<li>同一阻塞段填写同一区域ID，并仅在段首行写容量</li>"
            "<li>若需闸门等待，在 ‘起始需等区域ID’ 填下游区名</li>"
            "<li>最后一列『…』可自选颜色；未选自动配色</li>"
            "<li>顶部参数：车号数量 / 时间格刻度 / 等待分配方式</li>"
            "<li>填写完点击『生成并导出组合票』即可生成 Excel</li>"
            "</ol>"
        )
        QMessageBox.information(self, "帮助", msg)

    # ---------- 流程图弹窗 ----------
    def show_diagram(self):
        """弹出彩色流程图对话框"""
        steps = []
        for r in range(self.tbl.rowCount()):
            name = (self.tbl.item(r, 1).text().strip() if self.tbl.item(r, 1) else "")
            zone = (self.tbl.item(r, 5).text().strip() if self.tbl.item(r, 5) else "")
            gate = (self.tbl.item(r, 7).text().strip() if self.tbl.item(r, 7) else "")
            color = self.tbl.item(r, self.COL_COLOR).data(Qt.UserRole) or "#b0bec5"
            if name:
                steps.append({"row": r, "name": name, "zone": zone, "gate": gate, "color": color})
        if not steps:
            QMessageBox.information(self, "提示", "请先录入步骤")
       
